"""
Módulo de Exportação de Relatórios - PDF e Excel
Transmill v2.38.49

Endpoints:
- GET /api/exports/crm/excel - Exportar leads do CRM em Excel
- GET /api/exports/crm/pdf - Exportar relatório do CRM em PDF
- GET /api/exports/financeiro/excel - Exportar movimentações financeiras em Excel
- GET /api/exports/financeiro/pdf - Exportar relatório financeiro em PDF
"""

from fastapi import APIRouter, Depends, Query, Request
from fastapi.responses import StreamingResponse
from datetime import datetime, timedelta
from typing import Optional
import io
import logging

# Excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# PDF
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm, mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/exports", tags=["Exports"])

# Será injetado pelo server.py
_db = None
_get_current_user = None

def init_exports_routes(database, auth_dependency):
    """Inicializa as rotas de exportação com dependências"""
    global _db, _get_current_user
    _db = database
    _get_current_user = auth_dependency
    logger.info("✅ Módulo de exportação configurado")


# ============================================
# FUNÇÕES AUXILIARES
# ============================================

def create_excel_header_style():
    """Estilo para cabeçalhos do Excel"""
    return {
        'font': Font(bold=True, color='FFFFFF', size=11),
        'fill': PatternFill(start_color='0D47A1', end_color='0D47A1', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }

def create_excel_cell_style():
    """Estilo para células do Excel"""
    return {
        'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }

def format_date(dt):
    """Formata data para exibição"""
    if not dt:
        return '-'
    if isinstance(dt, str):
        try:
            dt = datetime.fromisoformat(dt.replace('Z', '+00:00'))
        except:
            return dt
    return dt.strftime('%d/%m/%Y %H:%M')

def format_currency(value):
    """Formata valor monetário"""
    try:
        return f"R$ {float(value):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    except:
        return 'R$ 0,00'


async def get_current_user_from_request(request: Request):
    """Obtém usuário atual a partir do request"""
    auth_header = request.headers.get('Authorization', '')
    if auth_header.startswith('Bearer '):
        token = auth_header.split(' ')[1]
        # Usar a função de autenticação injetada
        if _get_current_user:
            return await _get_current_user(token)
    return None


# ============================================
# EXPORTAÇÃO CRM - EXCEL
# ============================================

@router.get("/crm/excel")
async def export_crm_excel(
    request: Request,
    status: Optional[str] = Query(None, description="Filtrar por status"),
    data_inicio: Optional[str] = Query(None, description="Data início (YYYY-MM-DD)"),
    data_fim: Optional[str] = Query(None, description="Data fim (YYYY-MM-DD)")
):
    """
    Exporta leads do CRM Kanban em formato Excel
    """
    try:
        # Obter usuário atual
        current_user = await get_current_user_from_request(request)
        if not current_user:
            return {"success": False, "message": "Não autorizado"}
        
        logger.info(f"📊 [EXPORT] Gerando Excel CRM - User: {current_user.get('email')}")
        
        # Construir filtro
        filtro = {}
        
        # Filtro por hierarquia
        user_type = current_user.get('user_type')
        user_id = current_user.get('id')
        is_master = current_user.get('is_labelview_master', False)
        
        if not is_master:
            if user_type == 'labelview_consultor':
                filtro['consultor_id'] = user_id
            elif user_type == 'labelview_regional':
                filtro['regional_id'] = user_id
            elif user_type == 'labelview_unidade':
                filtro['unidade_id'] = user_id
        
        # Filtro por status
        if status:
            filtro['status'] = status
        
        # Filtro por data
        if data_inicio:
            try:
                dt_inicio = datetime.strptime(data_inicio, '%Y-%m-%d')
                filtro['created_at'] = {'$gte': dt_inicio}
            except:
                pass
        
        if data_fim:
            try:
                dt_fim = datetime.strptime(data_fim, '%Y-%m-%d') + timedelta(days=1)
                if 'created_at' in filtro:
                    filtro['created_at']['$lt'] = dt_fim
                else:
                    filtro['created_at'] = {'$lt': dt_fim}
            except:
                pass
        
        # Buscar leads
        leads = await _db.labelview_crm_leads.find(filtro).sort('created_at', -1).to_list(length=10000)
        
        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Leads CRM"
        
        # Cabeçalhos
        headers = ['ID', 'Nome', 'CPF', 'Email', 'Telefone', 'Status', 'Consultor', 'Data Criação', 'Observações']
        header_style = create_excel_header_style()
        cell_style = create_excel_cell_style()
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_style['font']
            cell.fill = header_style['fill']
            cell.alignment = header_style['alignment']
            cell.border = header_style['border']
        
        # Status labels
        status_labels = {
            'novo': 'Novo Lead',
            'interesse': 'Interesse',
            'negociacao': 'Negociação',
            'aguardando_docs': 'Aguardando Docs',
            'aprovado': 'Aprovado',
            'cancelado': 'Cancelado'
        }
        
        # Dados
        for row, lead in enumerate(leads, 2):
            data = [
                lead.get('id', '-')[:8] if lead.get('id') else '-',
                lead.get('nome', '-'),
                lead.get('cpf', '-'),
                lead.get('email', '-'),
                lead.get('telefone', '-'),
                status_labels.get(lead.get('status', ''), lead.get('status', '-')),
                lead.get('consultor_email', '-').split('@')[0] if lead.get('consultor_email') else '-',
                format_date(lead.get('created_at')),
                (lead.get('observacoes', '-')[:100] if lead.get('observacoes') else '-')
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.alignment = cell_style['alignment']
                cell.border = cell_style['border']
        
        # Ajustar largura das colunas
        column_widths = [12, 30, 15, 30, 18, 18, 20, 18, 40]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Congelar primeira linha
        ws.freeze_panes = 'A2'
        
        # Adicionar aba de resumo
        ws_resumo = wb.create_sheet("Resumo")
        ws_resumo['A1'] = "Resumo do CRM"
        ws_resumo['A1'].font = Font(bold=True, size=14)
        
        # Contagem por status
        ws_resumo['A3'] = "Status"
        ws_resumo['B3'] = "Quantidade"
        ws_resumo['A3'].font = Font(bold=True)
        ws_resumo['B3'].font = Font(bold=True)
        
        status_count = {}
        for lead in leads:
            s = lead.get('status', 'novo')
            status_count[s] = status_count.get(s, 0) + 1
        
        row = 4
        for s, count in status_count.items():
            ws_resumo.cell(row=row, column=1, value=status_labels.get(s, s))
            ws_resumo.cell(row=row, column=2, value=count)
            row += 1
        
        ws_resumo.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
        ws_resumo.cell(row=row, column=2, value=len(leads)).font = Font(bold=True)
        
        # Salvar em buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"crm_leads_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        logger.info(f"✅ [EXPORT] Excel CRM gerado: {len(leads)} leads")
        
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        logger.error(f"❌ [EXPORT] Erro ao gerar Excel CRM: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return {"success": False, "message": str(e)}


# ============================================
# EXPORTAÇÃO CRM - PDF
# ============================================

@router.get("/crm/pdf")
async def export_crm_pdf(request: Request):
    """
    Exporta relatório do CRM Kanban em formato PDF
    """
    try:
        # Obter usuário atual
        current_user = await get_current_user_from_request(request)
        if not current_user:
            return {"success": False, "message": "Não autorizado"}
        
        logger.info(f"📊 [EXPORT] Gerando PDF CRM - User: {current_user.get('email')}")
        
        # Construir filtro baseado em hierarquia
        filtro = {}
        user_type = current_user.get('user_type')
        user_id = current_user.get('id')
        is_master = current_user.get('is_labelview_master', False)
        
        if not is_master:
            if user_type == 'labelview_consultor':
                filtro['consultor_id'] = user_id
            elif user_type == 'labelview_regional':
                filtro['regional_id'] = user_id
            elif user_type == 'labelview_unidade':
                filtro['unidade_id'] = user_id
        
        # Buscar leads
        leads = await _db.labelview_crm_leads.find(filtro).sort('created_at', -1).to_list(length=10000)
        
        # Status labels
        status_labels = {
            'novo': 'Novo Lead',
            'interesse': 'Interesse',
            'negociacao': 'Negociação',
            'aguardando_docs': 'Aguardando Docs',
            'aprovado': 'Aprovado',
            'cancelado': 'Cancelado'
        }
        
        # Contagem por status
        status_count = {}
        for lead in leads:
            s = lead.get('status', 'novo')
            status_count[s] = status_count.get(s, 0) + 1
        
        # Criar PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=1.5*cm,
            leftMargin=1.5*cm,
            topMargin=1.5*cm,
            bottomMargin=1.5*cm
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Título
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#0D47A1'),
            spaceAfter=20,
            alignment=TA_CENTER
        )
        elements.append(Paragraph("Relatório CRM - Proteção Veicular", title_style))
        
        # Subtítulo com data
        subtitle_style = ParagraphStyle(
            'Subtitle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.gray,
            alignment=TA_CENTER,
            spaceAfter=30
        )
        elements.append(Paragraph(f"Gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}", subtitle_style))
        
        # Resumo
        section_style = ParagraphStyle(
            'Section',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#0D47A1'),
            spaceBefore=15,
            spaceAfter=10
        )
        elements.append(Paragraph("Resumo Geral", section_style))
        
        # Tabela de resumo
        resumo_data = [['Status', 'Quantidade', 'Percentual']]
        total = len(leads)
        for status_key in ['novo', 'interesse', 'negociacao', 'aguardando_docs', 'aprovado', 'cancelado']:
            count = status_count.get(status_key, 0)
            pct = f"{(count/total*100):.1f}%" if total > 0 else "0%"
            resumo_data.append([status_labels.get(status_key, status_key), str(count), pct])
        
        resumo_data.append(['TOTAL', str(total), '100%'])
        
        resumo_table = Table(resumo_data, colWidths=[6*cm, 3*cm, 3*cm])
        resumo_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0D47A1')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E3F2FD')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#BBDEFB')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F5F5F5')]),
        ]))
        elements.append(resumo_table)
        elements.append(Spacer(1, 20))
        
        # Métricas
        elements.append(Paragraph("Métricas de Conversão", section_style))
        
        aprovados = status_count.get('aprovado', 0)
        cancelados = status_count.get('cancelado', 0)
        taxa_conversao = f"{(aprovados/total*100):.1f}%" if total > 0 else "0%"
        taxa_cancelamento = f"{(cancelados/total*100):.1f}%" if total > 0 else "0%"
        
        metricas_data = [
            ['Métrica', 'Valor'],
            ['Total de Leads', str(total)],
            ['Leads Aprovados', str(aprovados)],
            ['Taxa de Conversão', taxa_conversao],
            ['Taxa de Cancelamento', taxa_cancelamento],
        ]
        
        metricas_table = Table(metricas_data, colWidths=[6*cm, 4*cm])
        metricas_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E7D32')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#C8E6C9')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#E8F5E9')]),
        ]))
        elements.append(metricas_table)
        elements.append(Spacer(1, 20))
        
        # Lista de leads (últimos 50)
        elements.append(Paragraph("Últimos Leads", section_style))
        
        leads_data = [['Nome', 'Status', 'Consultor', 'Data']]
        for lead in leads[:50]:
            leads_data.append([
                (lead.get('nome', '-')[:25] if lead.get('nome') else '-'),
                status_labels.get(lead.get('status', ''), '-'),
                ((lead.get('consultor_email', '-').split('@')[0] if lead.get('consultor_email') else '-')[:15]),
                format_date(lead.get('created_at')).split(' ')[0]
            ])
        
        leads_table = Table(leads_data, colWidths=[5*cm, 4*cm, 4*cm, 3*cm])
        leads_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0D47A1')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#BBDEFB')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
        ]))
        elements.append(leads_table)
        
        # Gerar PDF
        doc.build(elements)
        buffer.seek(0)
        
        filename = f"relatorio_crm_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        logger.info(f"✅ [EXPORT] PDF CRM gerado: {len(leads)} leads")
        
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        logger.error(f"❌ [EXPORT] Erro ao gerar PDF CRM: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return {"success": False, "message": str(e)}


# ============================================
# EXPORTAÇÃO FINANCEIRO - EXCEL
# ============================================

@router.get("/financeiro/excel")
async def export_financeiro_excel(
    request: Request,
    tipo: Optional[str] = Query(None, description="entrada, saida ou todos"),
    data_inicio: Optional[str] = Query(None),
    data_fim: Optional[str] = Query(None)
):
    """
    Exporta movimentações financeiras em Excel
    """
    try:
        # Obter usuário atual
        current_user = await get_current_user_from_request(request)
        if not current_user:
            return {"success": False, "message": "Não autorizado"}
        
        logger.info(f"📊 [EXPORT] Gerando Excel Financeiro - User: {current_user.get('email')}")
        
        # Construir filtro
        filtro = {}
        
        if tipo and tipo != 'todos':
            filtro['tipo'] = tipo
        
        if data_inicio:
            try:
                dt_inicio = datetime.strptime(data_inicio, '%Y-%m-%d')
                filtro['data'] = {'$gte': dt_inicio}
            except:
                pass
        
        if data_fim:
            try:
                dt_fim = datetime.strptime(data_fim, '%Y-%m-%d') + timedelta(days=1)
                if 'data' in filtro:
                    filtro['data']['$lt'] = dt_fim
                else:
                    filtro['data'] = {'$lt': dt_fim}
            except:
                pass
        
        # Buscar movimentações
        movimentacoes = await _db.movimentacoes_bolsao.find(filtro).sort('data', -1).to_list(length=10000)
        
        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Movimentações"
        
        # Cabeçalhos
        headers = ['Data', 'Tipo', 'Valor', 'Descrição', 'Franquia', 'Origem']
        header_style = create_excel_header_style()
        cell_style = create_excel_cell_style()
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_style['font']
            cell.fill = header_style['fill']
            cell.alignment = header_style['alignment']
            cell.border = header_style['border']
        
        # Totais
        total_entrada = 0
        total_saida = 0
        
        # Dados
        for row, mov in enumerate(movimentacoes, 2):
            tipo_mov = mov.get('tipo', '-')
            valor = float(mov.get('valor', 0))
            
            if tipo_mov == 'entrada':
                total_entrada += valor
            elif tipo_mov == 'saida':
                total_saida += valor
            
            data = [
                format_date(mov.get('data')),
                'Entrada' if tipo_mov == 'entrada' else 'Saída',
                format_currency(valor),
                (mov.get('descricao', '-')[:50] if mov.get('descricao') else '-'),
                mov.get('franquia_nome', '-'),
                mov.get('origem', '-')
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.alignment = cell_style['alignment']
                cell.border = cell_style['border']
                
                # Cor verde para entrada, vermelho para saída
                if col == 2:
                    if tipo_mov == 'entrada':
                        cell.font = Font(color='2E7D32')
                    else:
                        cell.font = Font(color='C62828')
        
        # Ajustar largura das colunas
        column_widths = [18, 12, 18, 40, 25, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Congelar primeira linha
        ws.freeze_panes = 'A2'
        
        # Adicionar aba de resumo
        ws_resumo = wb.create_sheet("Resumo")
        ws_resumo['A1'] = "Resumo Financeiro"
        ws_resumo['A1'].font = Font(bold=True, size=14)
        
        ws_resumo['A3'] = "Total Entradas:"
        ws_resumo['B3'] = format_currency(total_entrada)
        ws_resumo['B3'].font = Font(color='2E7D32', bold=True)
        
        ws_resumo['A4'] = "Total Saídas:"
        ws_resumo['B4'] = format_currency(total_saida)
        ws_resumo['B4'].font = Font(color='C62828', bold=True)
        
        ws_resumo['A5'] = "Saldo:"
        ws_resumo['B5'] = format_currency(total_entrada - total_saida)
        ws_resumo['B5'].font = Font(bold=True, size=12)
        
        ws_resumo['A7'] = "Período:"
        ws_resumo['B7'] = f"{data_inicio or 'Início'} a {data_fim or 'Hoje'}"
        
        ws_resumo['A8'] = "Total de Movimentações:"
        ws_resumo['B8'] = str(len(movimentacoes))
        
        # Salvar em buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"financeiro_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        logger.info(f"✅ [EXPORT] Excel Financeiro gerado: {len(movimentacoes)} movimentações")
        
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        logger.error(f"❌ [EXPORT] Erro ao gerar Excel Financeiro: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return {"success": False, "message": str(e)}


# ============================================
# EXPORTAÇÃO FINANCEIRO - PDF
# ============================================

@router.get("/financeiro/pdf")
async def export_financeiro_pdf(request: Request):
    """
    Exporta relatório financeiro em PDF
    """
    try:
        # Obter usuário atual
        current_user = await get_current_user_from_request(request)
        if not current_user:
            return {"success": False, "message": "Não autorizado"}
        
        logger.info(f"📊 [EXPORT] Gerando PDF Financeiro - User: {current_user.get('email')}")
        
        # Buscar movimentações (últimos 30 dias)
        data_30_dias = datetime.utcnow() - timedelta(days=30)
        movimentacoes = await _db.movimentacoes_bolsao.find({
            'data': {'$gte': data_30_dias}
        }).sort('data', -1).to_list(length=1000)
        
        # Calcular totais
        total_entrada = sum(float(m.get('valor', 0)) for m in movimentacoes if m.get('tipo') == 'entrada')
        total_saida = sum(float(m.get('valor', 0)) for m in movimentacoes if m.get('tipo') == 'saida')
        saldo = total_entrada - total_saida
        
        # Criar PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=1.5*cm,
            leftMargin=1.5*cm,
            topMargin=1.5*cm,
            bottomMargin=1.5*cm
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Título
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#2E7D32'),
            spaceAfter=20,
            alignment=TA_CENTER
        )
        elements.append(Paragraph("Relatório Financeiro - Conta Bolsão", title_style))
        
        # Subtítulo
        subtitle_style = ParagraphStyle(
            'Subtitle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.gray,
            alignment=TA_CENTER,
            spaceAfter=30
        )
        elements.append(Paragraph(f"Últimos 30 dias - Gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}", subtitle_style))
        
        # Resumo
        section_style = ParagraphStyle(
            'Section',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#2E7D32'),
            spaceBefore=15,
            spaceAfter=10
        )
        elements.append(Paragraph("Resumo Financeiro", section_style))
        
        # Tabela de resumo
        resumo_data = [
            ['Descrição', 'Valor'],
            ['Total de Entradas', format_currency(total_entrada)],
            ['Total de Saídas', format_currency(total_saida)],
            ['Saldo do Período', format_currency(saldo)],
            ['Qtd. Movimentações', str(len(movimentacoes))],
        ]
        
        resumo_table = Table(resumo_data, colWidths=[8*cm, 6*cm])
        resumo_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E7D32')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#C8E6C9')),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E8F5E9')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            # Cor verde para entradas
            ('TEXTCOLOR', (1, 1), (1, 1), colors.HexColor('#2E7D32')),
            # Cor vermelha para saídas
            ('TEXTCOLOR', (1, 2), (1, 2), colors.HexColor('#C62828')),
        ]))
        elements.append(resumo_table)
        elements.append(Spacer(1, 20))
        
        # Últimas movimentações
        elements.append(Paragraph("Últimas Movimentações", section_style))
        
        mov_data = [['Data', 'Tipo', 'Valor', 'Descrição']]
        for mov in movimentacoes[:30]:
            tipo = 'Entrada' if mov.get('tipo') == 'entrada' else 'Saída'
            mov_data.append([
                format_date(mov.get('data')).split(' ')[0],
                tipo,
                format_currency(mov.get('valor', 0)),
                (mov.get('descricao', '-')[:30] if mov.get('descricao') else '-')
            ])
        
        mov_table = Table(mov_data, colWidths=[3*cm, 3*cm, 4*cm, 6*cm])
        mov_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E7D32')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#C8E6C9')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
        ]))
        elements.append(mov_table)
        
        # Gerar PDF
        doc.build(elements)
        buffer.seek(0)
        
        filename = f"relatorio_financeiro_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        logger.info(f"✅ [EXPORT] PDF Financeiro gerado: {len(movimentacoes)} movimentações")
        
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        logger.error(f"❌ [EXPORT] Erro ao gerar PDF Financeiro: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return {"success": False, "message": str(e)}
